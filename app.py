import os
from flask import Flask, flash, request, redirect, url_for
from werkzeug.utils import secure_filename
from flask import send_from_directory, render_template
from utils import process_file, allowed_file
import googlemaps
from haversine import haversine
from config import UPLOAD_FOLDER
import openpyxl
from db.utils import Database


app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


@app.route('/resultado')
def resultado():
    locations = []
    sorocaba = None
    arquivo = os.path.join(app.config["UPLOAD_FOLDER"], 'retorno.xlsx')
    if os.path.exists(arquivo):
        wb_obj = openpyxl.load_workbook(arquivo)
        sheet_obj = wb_obj.active
        for x in sheet_obj.iter_rows(min_row=2, max_row=sheet_obj.max_row, min_col=1, max_col=6):
            location = '{},{}'.format(x[1].value, x[0].value)
            distancia_sorocaba = '{} Km'.format(round(x[3].value / 1000, 2))
            info = x[5].value
            locations.append({'location': location,
                              'code': None,
                              'distancia': distancia_sorocaba,
                              'info': info
                              })

        sorocaba = geolocate.geocode(address='Sorocaba,SP', language='pt_BR')
        sorocaba = sorocaba[0]['geometry']['location']

        lat_sor = sorocaba['lat']
        lng_sor = sorocaba['lng']

        sor_coord = (lat_sor, lng_sor)

        for y, x in enumerate(locations):
            location = geolocate.geocode(address=x['location'], language='pt_BR')
            locations[y]['code'] = location[0]['geometry']['location']
            lat = location[0]['geometry']['location']['lat']
            lng = location[0]['geometry']['location']['lng']
            loc = (lat, lng)
            locations[y]['dist_rad'] = round(haversine(sor_coord, loc), 2)

    return render_template('resultado.html', sorocaba=sorocaba, locacoes=locations)


@app.route('/uploads')
def download_file():
    return send_from_directory(app.config["UPLOAD_FOLDER"], 'retorno.xlsx')

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        gmaps_distance = googlemaps.Client(key=os.environ['DISTANCE_API_KEY'])
        geolocate_maps = googlemaps.Client(key=os.environ['GEOLOCATE_API_KEY'])
        if 'file' not in request.files:
            flash('No file part')
            return redirect(request.url)
        file = request.files['file']
        if file.filename == '':
            flash('No selected file')
            return redirect(request.url)
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            arquivo = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(arquivo)
            process_file(arquivo,gmaps_distance,geolocate_maps)
            return redirect(url_for('resultado'))
    return render_template('base.html')


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=int(os.environ.get("PORT", 8080)))
