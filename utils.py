import os
import openpyxl
from db.models import City
from db.utils import Database
from config import ALLOWED_EXTENSIONS


def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]



def get_data_from_sheet(sheet):
    wb_obj = openpyxl.load_workbook(sheet)
    sheet_obj = wb_obj.active
    linhas = sheet_obj.max_row

    linha_ini = 2
    destinos = []
    while linha_ini <= linhas:
        linha_ini += 1
        destinos.append(
            {
                'city':sheet_obj.cell(linha_ini, 2).value,
                'state':sheet_obj.cell(linha_ini, 1).value
            }
        )
    return destinos

def find_lat_lng(gmaps_locate,adderss):
    coords = gmaps_locate.geocode(address=adderss, language='pt_BR')
    return coords[0]['geometry']['location']['lat'],coords[0]['geometry']['location']['lng']



def process_file(arquivo,gmaps_distance,gmaps_locate):
    origin = '{},{}'.format(os.environ['LOCATION_ORIGIN_CITY'],os.environ['LOCATION_ORIGIN_STATE'])
    destinos = get_data_from_sheet(arquivo)

    for destino in destinos:
        if destino['city'] and destino['state']:
            city = City()
            city.name = destino['city']
            city.state = destino['state']
            driver = Database().driver
            busca = City.search(city,driver)
            if not busca:
                address = '{},{}'.format(destino['city'],destino['state'])
                coords = find_lat_lng(gmaps_locate,address)
                n_city = City(
                    name=destino['city'],
                    state=destino['state'],
                    latitude=coords[0],
                    longitude=coords[1]
                ).save()
                print(n_city)


    chunks_s = list(chunks(destinos, 20))

    results = []
    for x in chunks_s:
        directions_result = gmaps_distance.distance_matrix(
            origin,
            x,
            mode="driving",
            language='pt_BR',
            units="metric")
        results.append({'destinos': x, 'retornos': directions_result})

    linha_ini = 2
    sheet_obj.cell(1, 4).value = 'Distância em Metros'
    sheet_obj.cell(1, 5).value = 'Tempo de Viagem'
    for x in results:
        x_retornos = x['retornos']['rows'][0]['elements']
        for k in x_retornos:
            sheet_obj.cell(linha_ini, 4).value = k['distance']['value']
            sheet_obj.cell(linha_ini, 5).value = k['duration']['text']
            linha_ini += 1

    wb_obj.save(os.path.join('arquivos', 'retorno.xlsx'))