import os
import googlemaps
import openpyxl
gmaps = googlemaps.Client(key='AIzaSyCXi-8bUjZ2lEyac9VbHkPHCjrCF-kLFCE')


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

def process_file(arquivo):

    origin = "Sorocaba,SP"

    wb_obj = openpyxl.load_workbook(arquivo)

    sheet_obj = wb_obj.active

    linhas = sheet_obj.max_row

    linha_ini = 2
    destinos = []
    while linha_ini<=linhas:
        destino = '{},{}'.format(sheet_obj.cell(linha_ini,2).value,sheet_obj.cell(linha_ini,1).value)
        linha_ini+=1
        destinos.append(destino)

    chunks_s = list(chunks(destinos,20))

    results = []
    for x in chunks_s:

        directions_result = gmaps.distance_matrix(
                                           origin,
                                             x,
                                             mode="driving",
                                             language='pt_BR',
                                            units="metric")
        results.append({'destinos':x,'retornos':directions_result})

    linha_ini = 2
    sheet_obj.cell(1,4).value='Distância em Metros'
    sheet_obj.cell(1,5).value='Tempo de Viagem'
    for x in results:
        x_retornos = x['retornos']['rows'][0]['elements']
        for k in x_retornos:
            sheet_obj.cell(linha_ini,4).value=k['distance']['value']
            sheet_obj.cell(linha_ini, 5).value = k['duration']['text']
            linha_ini+=1


    wb_obj.save(os.path.join('arquivos','retorno.xlsx'))